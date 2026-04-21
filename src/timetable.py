"""Timetable module - matches subjects and calculates due dates."""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


# Mapping from email subject names to timetable names
SUBJECT_MAPPING = {
    # Teacher name in email : Subject in timetable
    "Гонтковська": "Мистецтво",
    "Сидоренко": "Українська мова",
    "Колесник": "Англійська мова",
    "Шіхтар": "Фізика",
    "Ланій": "Біологія",
    "Ткаченко": "Історія України",
    "Шевченко": "Геометрія",
    "Бондаренко": "Алгебра",
    "Максимчук": "Хімія",
    "Старовір": "Географія",
    "Павленко": "Інформатика",
    # Direct subject name mappings (UPPERCASE in email)
    "МИСТЕЦТВО": "Мистецтво",
    "БІОЛОГІЯ": "Біологія",
    "ФІЗИКА": "Фізика",
    "ХІМІЯ": "Хімія",
    "ГЕОГРАФІЯ": "Географія",
    "ІНФОРМАТИКА": "Інформатика",
    "АЛГЕБРА": "Алгебра",
    "ГЕОМЕТРІЯ": "Геометрія",
    "ІСТОРІЯ": "Історія України",
    "УКРАЇНСЬКА": "Українська мова",
    "АНГЛІЙСЬКА": "Англійська мова",
}


class Timetable:
    """Manages class timetable."""

    def __init__(self, timetable_file: str = "timetable.xlsx"):
        self.timetable_file = Path(timetable_file)
        self._workbook = None
        self._worksheet = None
        self._load()

    def _load(self):
        """Load timetable from Excel."""
        if not self.timetable_file.exists():
            logger.warning(f"Timetable file not found: {self.timetable_file}")
            return

        try:
            self._workbook = openpyxl.load_workbook(self.timetable_file)
            self._worksheet = self._workbook.active
            logger.info(f"Loaded timetable from {self.timetable_file}")
        except Exception as e:
            logger.error(f"Failed to load timetable: {e}")

    def get_subject_from_teacher(self, teacher_name: str) -> Optional[str]:
        """Get subject name from teacher name."""
        if not teacher_name:
            return None

        # Try exact match first
        if teacher_name in SUBJECT_MAPPING:
            return SUBJECT_MAPPING[teacher_name]

        # Try partial match (teacher last name)
        teacher_lower = teacher_name.lower()
        for name, subject in SUBJECT_MAPPING.items():
            if name.lower() in teacher_lower:
                return subject

        return None

    def normalize_subject(self, subject: str) -> str:
        """Normalize subject name to timetable format."""
        if not subject:
            return subject

        # Check mapping
        if subject.upper() in {k.upper(): v for k, v in SUBJECT_MAPPING.items()}:
            for k, v in SUBJECT_MAPPING.items():
                if k.upper() == subject.upper():
                    return v

        # If not found in mapping, return as-is
        return subject

    def find_next_lesson(self, subject: str, from_date: datetime) -> Optional[datetime]:
        """Find next lesson for subject after from_date."""
        if not self._worksheet:
            return None

        if not subject:
            return None

        # Normalize subject - uppercase for comparison
        subject_upper = subject.upper()

        # Try to find subject in column 3 (subject column)
        best_date = None
        from_date_only = from_date.date()

        for row in range(2, self._worksheet.max_row + 1):
            cell_subject = self._worksheet.cell(row=row, column=3).value
            if not cell_subject:
                continue

            # Check if this row has our subject (case-insensitive)
            if subject_upper in str(cell_subject).upper():
                # Get day - column 1 has time, row above has day name
                # Look up for day name in previous rows
                day_name = None
                for check_row in range(row - 1, 0, -1):
                    val = self._worksheet.cell(row=check_row, column=1).value
                    if val and str(val).lower() in ['понеділок', 'вівторок', 'середа', 'четвер', 'пятниця', 'субота']:
                        day_name = val
                        break

                if not day_name:
                    continue

                # Parse day
                lesson_day = self._parse_day(day_name, from_date)
                if lesson_day and lesson_day >= from_date_only:
                    if best_date is None or lesson_day < best_date:
                        best_date = lesson_day

        if best_date:
            return datetime.combine(best_date, datetime.min.time().replace(hour=23, minute=59))

        return None

    def _parse_day(self, day_str: str, from_date: datetime = None) -> Optional[datetime]:
        """Parse day string like 'Понеділок', 'Вівторок', etc."""
        if not day_str:
            return None

        if from_date is None:
            from_date = datetime.now()

        days_map = {
            "понеділок": 0,
            "вівторок": 1,
            "середа": 2,
            "четвер": 3,
            "пятниця": 4,
            "субота": 5,
        }

        day_lower = str(day_str).lower().strip()
        if day_lower in days_map:
            target_weekday = days_map[day_lower]

            # Find next occurrence of this weekday after from_date
            today = from_date.date()
            days_until = (target_weekday - today.weekday()) % 7
            if days_until == 0:
                # If same day, use next week's occurrence
                days_until = 7
            elif days_until < 0:
                days_until = 7 + days_until

            return today + timedelta(days=days_until)

        return None