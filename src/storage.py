"""Assignment storage in Excel file."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .parser import ClassroomAssignment

logger = logging.getLogger(__name__)


class AssignmentStorage:
    """Manages assignments in Excel file."""

    def __init__(self, storage_file: str = "assignments.xlsx", timetable_file: str = "timetable.xlsx"):
        self.storage_file = Path(storage_file)
        self.timetable_file = timetable_file
        self._workbook = None
        self._worksheet = None
        self._load_or_create()

    def _load_or_create(self):
        """Load existing workbook or create new one."""
        if self.storage_file.exists():
            try:
                self._workbook = openpyxl.load_workbook(self.storage_file)
                if "Assignments" in self._workbook.sheetnames:
                    self._worksheet = self._workbook["Assignments"]
                    # Check if headers need update
                    if self._worksheet.cell(1, 1).value != "ID":
                        self._setup_headers()
                else:
                    self._worksheet = self._workbook.create_sheet("Assignments")
                    self._setup_headers()
            except Exception as e:
                logger.warning(f"Failed to load workbook: {e}, creating new")
                self._workbook = openpyxl.Workbook()
                self._worksheet = self._workbook.active
                self._worksheet.title = "Assignments"
                self._setup_headers()
        else:
            self._workbook = openpyxl.Workbook()
            self._worksheet = self._workbook.active
            self._worksheet.title = "Assignments"
            self._setup_headers()

    def _setup_headers(self):
        """Setup table headers."""
        headers = [
            "ID",
            "Предмет",
            "Завдання",
            "Повний текст",
            "Вчитель",
            "Отримано",
            "Срок",
            "Час строку",
            "Посилання",
            "Email ID",
            "Виконано",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = self._worksheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Set column widths
        widths = [10, 25, 50, 80, 30, 18, 18, 12, 50, 20, 12]
        for col, width in enumerate(widths, start=1):
            self._worksheet.column_dimensions[get_column_letter(col)].width = width

        # Set wrap text for all columns
        for row in range(1, self._worksheet.max_row + 1):
            for col in range(1, 12):
                cell = self._worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Set row height for data rows
            if row > 1:
                self._worksheet.row_dimensions[row].height = 60

    def _generate_id(self) -> str:
        existing_ids = set()
        for row in range(2, self._worksheet.max_row + 1):
            cell_value = self._worksheet.cell(row=row, column=1).value
            if cell_value:
                existing_ids.add(cell_value)

        base_id = datetime.now().strftime("%Y%m%d%H%M%S")
        counter = 1
        while f"{base_id}_{counter}" in existing_ids:
            counter += 1

        return f"{base_id}_{counter}"

    def add_assignment(self, assignment: ClassroomAssignment) -> str:
        """Add new assignment to storage. Returns assignment ID."""
        assignment_id = self._generate_id()

        # Check if assignment already exists (by email_id)
        existing_row = self._find_by_email_id(assignment.email_id)
        if existing_row:
            logger.info(f"Assignment {assignment.email_id} already exists, updating")
            self._update_assignment(existing_row, assignment)
            return self._worksheet.cell(row=existing_row, column=1).value

        # Calculate due date if not provided (use timetable)
        due_date = assignment.due_date

        # If no due date in email, calculate from timetable or fallback to 2 days
        if not due_date:
            # Try to get subject
            subject = assignment.course_name
            if not subject or subject == "Unknown":
                if assignment.teacher_name:
                    from .subject_mappings import TEACHER_SUBJECTS
                    for teacher_key, subjects_list in TEACHER_SUBJECTS.items():
                        if teacher_key.split()[0].lower() in assignment.teacher_name.lower():
                            subject = subjects_list[0]
                            break

            # Try timetable
            if subject and subject != "Unknown":
                try:
                    from .timetable import Timetable
                    timetable = Timetable(self.timetable_file)
                    next_lesson = timetable.find_next_lesson(subject, assignment.received_date)
                    if next_lesson:
                        due_date = next_lesson
                        logger.info(f"Calculated due date from timetable: {due_date}")
                except Exception as e:
                    logger.warning(f"Timetable error: {e}")

            # Fallback: 2 days if no timetable match
            if not due_date:
                from datetime import timedelta
                due_date = assignment.received_date + timedelta(days=2)
                logger.info(f"Using fallback due date: {due_date}")

        # Add new row
        row = self._worksheet.max_row + 1

        self._worksheet.cell(row=row, column=1).value = assignment_id
        self._worksheet.cell(row=row, column=2).value = assignment.course_name
        self._worksheet.cell(row=row, column=3).value = assignment.assignment_title
        self._worksheet.cell(row=row, column=4).value = assignment.assignment_text
        self._worksheet.cell(row=row, column=5).value = assignment.teacher_name
        self._worksheet.cell(row=row, column=6).value = assignment.received_date.strftime("%Y-%m-%d %H:%M")
        self._worksheet.cell(row=row, column=7).value = (
            due_date.strftime("%Y-%m-%d") if due_date else ""
        )
        self._worksheet.cell(row=row, column=8).value = assignment.due_time or ""
        self._worksheet.cell(row=row, column=9).value = assignment.link
        self._worksheet.cell(row=row, column=10).value = assignment.email_id
        self._worksheet.cell(row=row, column=11).value = "Ні"

        # Format due date if present
        if due_date:
            due_cell = self._worksheet.cell(row=row, column=6)
            due_cell.alignment = Alignment(horizontal="center")

        logger.info(f"Added assignment: {assignment_id} - {assignment.course_name}")
        self.save()

        return assignment_id

    def _find_by_email_id(self, email_id: str) -> Optional[int]:
        """Find row by email ID."""
        for row in range(2, self._worksheet.max_row + 1):
            if self._worksheet.cell(row=row, column=10).value == email_id:
                return row
        return None

    def _update_assignment(self, row: int, assignment: ClassroomAssignment):
        """Update existing assignment."""
        self._worksheet.cell(row=row, column=2).value = assignment.course_name
        self._worksheet.cell(row=row, column=3).value = assignment.assignment_title
        self._worksheet.cell(row=row, column=4).value = assignment.assignment_text
        self._worksheet.cell(row=row, column=5).value = assignment.teacher_name
        self._worksheet.cell(row=row, column=6).value = assignment.received_date.strftime("%Y-%m-%d %H:%M")
        if assignment.due_date:
            self._worksheet.cell(row=row, column=7).value = assignment.due_date.strftime("%Y-%m-%d")
        if assignment.due_time:
            self._worksheet.cell(row=row, column=8).value = assignment.due_time
        if assignment.link:
            self._worksheet.cell(row=row, column=9).value = assignment.link

        self.save()

    def save(self):
        """Save workbook to file."""
        try:
            self.storage_file.parent.mkdir(parents=True, exist_ok=True)
            self._workbook.save(self.storage_file)
            logger.debug(f"Saved to {self.storage_file}")
        except Exception as e:
            logger.error(f"Failed to save: {e}")

    def get_assignments(self, completed: Optional[bool] = None) -> list[dict]:
        """Get all assignments as list of dicts."""
        assignments = []

        for row in range(2, self._worksheet.max_row + 1):
            is_completed = self._worksheet.cell(row=row, column=11).value == "Так"

            if completed is not None and is_completed != completed:
                continue

            assignment = {
                "id": self._worksheet.cell(row=row, column=1).value,
                "course": self._worksheet.cell(row=row, column=2).value,
                "title": self._worksheet.cell(row=row, column=3).value,
                "full_text": self._worksheet.cell(row=row, column=4).value,
                "teacher": self._worksheet.cell(row=row, column=5).value,
                "received": self._worksheet.cell(row=row, column=6).value,
                "due_date": self._worksheet.cell(row=row, column=7).value,
                "due_time": self._worksheet.cell(row=row, column=8).value,
                "link": self._worksheet.cell(row=row, column=9).value,
                "completed": is_completed,
            }
            assignments.append(assignment)

        return assignments

    def get_assignments_for_date(self, target_date: str) -> list[dict]:
        """Get assignments due on specific date (YYYY-MM-DD format)."""
        all_assignments = self.get_assignments(completed=False)
        return [a for a in all_assignments if a.get("due_date") == target_date]

    def mark_completed(self, assignment_id: str):
        """Mark assignment as completed."""
        for row in range(2, self._worksheet.max_row + 1):
            if self._worksheet.cell(row=row, column=1).value == assignment_id:
                self._worksheet.cell(row=row, column=10).value = "Так"
                self.save()
                logger.info(f"Marked assignment {assignment_id} as completed")
                return True
        return False

    def delete_assignment(self, assignment_id: str):
        """Delete assignment by ID."""
        for row in range(2, self._worksheet.max_row + 1):
            if self._worksheet.cell(row=row, column=1).value == assignment_id:
                self._worksheet.delete_rows(row)
                self.save()
                logger.info(f"Deleted assignment {assignment_id}")
                return True
        return False